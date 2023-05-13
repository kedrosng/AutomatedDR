import re
import os
import csv
import glob
from datetime import datetime
from tabulate import tabulate
from collections import namedtuple
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class Drawing:
    def __init__(self, drawing_number, title, revision, date, location_code, drawing_type, submission_date=None, submission_ref=None):
        self.drawing_number = drawing_number
        self.title = title
        self.revision = revision
        self.date = date
        self.location_code = location_code
        self.drawing_type = drawing_type
        self.submission_date = submission_date
        self.submission_ref = submission_ref

    def __str__(self):
        return (f"{self.drawing_number}: {self.title} "
                f"(Location Code: {self.location_code}, Drawing Type: {self.drawing_type}, "
                f"Revision: {self.revision}, Date: {self.date}, "
                f"Submission Date: {self.submission_date}, Submission Ref: {self.submission_ref})")


def process_all_drawings(base_folder, drawings_register):
    # Prepare the log file
    with open("log.txt", "w") as log_file:
        # Recursively process all folders and subfolders in the base_folder
        for root, dirs, files in os.walk(base_folder):
            for drawing_type_folder in dirs:
                drawing_type_path = os.path.join(root, drawing_type_folder)

                print(f"Processing drawing type folder: {drawing_type_path}")

                submission_date = datetime.fromtimestamp(os.path.getmtime(drawing_type_path)).strftime('%Y-%m-%d')

                # Find all the drawing files in the drawing_type_path
                drawing_files = glob.glob(os.path.join(drawing_type_path, "*"))

                # Filter out the files that do not match the file name format criteria
                valid_files = [drawing_file for drawing_file in drawing_files
                                if re.search(r"HKT2_BYME_SDWG_([A-Za-z0-9]+)[_-]([A-Za-z0-9]+)[_-](\d+)[_-](\w)", os.path.basename(drawing_file))]

                invalid_files = set(drawing_files) - set(valid_files)

                for invalid_file in invalid_files:
                    log_file.write(f"Invalid file format: {invalid_file}\n")

                print(f"Valid files found: {len(valid_files)}")
                print(f"Invalid files found: {len(invalid_files)}")

                if valid_files:
                    update_drawings_in_batch(drawings_register, valid_files, submission_date, drawing_type_folder)


def parse_filename(filename):
    pattern = r"HKT2_BYME_SDWG_([A-Za-z0-9]+)[_-]([A-Za-z0-9]+)[_-](\d+)[_-]([A-Za-z0-9]*)\.pdf"
    match = re.search(pattern, filename)
    if match:
        location_code, drawing_type, drawing_number, revision = match.groups()
        return {
            "location_code": location_code,
            "drawing_type": drawing_type,
            "drawing_number": int(drawing_number),
            "revision": revision
        }
    return None


def update_drawings_in_batch(drawings, filepaths, submission_date, submission_ref):
    updated_drawings = 0
    for filepath in filepaths:
        drawing_info = parse_filename(os.path.basename(filepath))
        if drawing_info is None:
            print(f"Error: could not extract drawing info from file {filepath}")
            continue

        # Find the existing drawing with the same drawing number, location code, drawing type, and revision
        existing_drawing = None
        for drawing in drawings:
            if (drawing.drawing_number == drawing_info["drawing_number"]
                    and drawing.location_code == drawing_info["location_code"]
                    and drawing.drawing_type == drawing_info["drawing_type"]
                    and drawing.revision == drawing_info["revision"]):
                existing_drawing = drawing
                break

        title = os.path.splitext(os.path.basename(filepath))[0]

        if existing_drawing:
            existing_drawing.title = title
            existing_drawing.date = submission_date
            existing_drawing.submission_date = submission_date
            existing_drawing.submission_ref = submission_ref
            updated_drawings += 1
        else:
            # Create a new drawing object and add it to the drawings register
            new_drawing = Drawing(
                drawing_number=drawing_info["drawing_number"],
                title=title,
                revision=drawing_info["revision"],
                date=submission_date,
                location_code=drawing_info["location_code"],
                drawing_type=drawing_info["drawing_type"],
                submission_date=submission_date,
                submission_ref=submission_ref,
            )
            drawings.append(new_drawing)
            updated_drawings += 1

            print(f"Updated {updated_drawings} drawings in the batch")


def save_drawings_to_csv(drawings, output_folder):
    drawings_by_type = {}

    # Group drawings by "location_code_drawing type" combination
    for drawing in drawings:
        key = f"{drawing.location_code}_{drawing.drawing_type}"
        if key not in drawings_by_type:
            drawings_by_type[key] = []
        drawings_by_type[key].append(drawing)

    # Create output folder if it does not exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Write separate CSV files for each "location_code_drawing type" combination
    for key, drawings in drawings_by_type.items():
        output_file = os.path.join(output_folder, f"{key}.csv")
        with open(output_file, "w", newline="") as csvfile:
            fieldnames = ["drawing_number", "title", "revision", "date", "location_code", "drawing_type", "submission_date", "submission_ref"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for drawing in drawings:
                writer.writerow({
                    "drawing_number": f"{drawing.drawing_number:06d}",
                    "title": drawing.title,
                    "revision": drawing.revision,
                    "date": drawing.date,
                    "location_code": drawing.location_code,
                    "drawing_type": drawing.drawing_type,
                    "submission_date": drawing.submission_date,
                    "submission_ref": drawing.submission_ref,
                })
def create_timestamped_file():
    # Get the current timestamp and format it as a string
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # Create the text file with the timestamp in its name
    filename = f"{timestamp}.txt"
    with open(filename, "w") as file:
        file.write(f"Script run at: {timestamp}\n")
def save_drawings_to_excel(drawings, output_file):
    drawings_by_type = {}

    # Group drawings by "location_code_drawing type" combination
    for drawing in drawings:
        key = f"{drawing.location_code}_{drawing.drawing_type}"
        if key not in drawings_by_type:
            drawings_by_type[key] = []
        drawings_by_type[key].append(drawing)

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    workbook.remove(sheet)
  

    # Write separate sheets for each "location_code_drawing type" combination
    for key, drawings in drawings_by_type.items():
        sheet = workbook.create_sheet(title=key)
        fieldnames = ["drawing_number", "title", "revision", "date", "location_code", "drawing_type", "submission_date", "submission_ref"]

        # Write header row and style it
        for col_num, fieldname in enumerate(fieldnames, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = fieldname
            cell.font = Font(bold=True)

        # Write data rows
        for row_num, drawing in enumerate(drawings, start=2):
            sheet.cell(row=row_num, column=1, value=drawing.title)
    
            # Use row_num instead of 2 in the formula
            title_formula = f"=VLOOKUP(LEFT(A{row_num},29),[aconex_export.xls]Docs!$B$13:$E$9000,4,FALSE)"
    
            sheet.cell(row=row_num, column=2, value=title_formula)
            sheet.cell(row=row_num, column=3, value=drawing.revision)
            sheet.cell(row=row_num, column=4, value=drawing.date)
            sheet.cell(row=row_num, column=5, value=drawing.location_code)
            sheet.cell(row=row_num, column=6, value=drawing.drawing_type)
            sheet.cell(row=row_num, column=7, value=drawing.submission_date)
            sheet.cell(row=row_num, column=8, value=drawing.submission_ref)

        # Adjust column widths
        for col_num in range(1, len(fieldnames) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].auto_size = True

    # Save the Excel workbook to the output file
    workbook.save(output_file)
def main():
    # Create a list to store all the drawings
    drawings_register = []

    # Prompt the user to select a folder
    root = tk.Tk()
    root.withdraw()
    base_folder = filedialog.askdirectory(title="Select a folder containing the drawing files")

    if not base_folder:
        print("No folder selected. Exiting.")
        return

    # Process all drawings in the selected folder and its subfolders
    process_all_drawings(base_folder, drawings_register)

    # Ask the user if they want to save the drawings register as CSV files
    save_as_csv = input("Do you want to save the updated drawings register as CSV files? (y/n): ").lower()

    if save_as_csv == 'y':
        # Save the updated drawings register to separate CSV files based on "location_code_drawing type"
        output_folder = "output_drawings"
        save_drawings_to_csv(drawings_register, output_folder)
        print("Saved updated drawings register to CSV files in:", output_folder)
    else:
        print("Skipping saving updated drawings register to CSV files.")

    # Save the updated drawings register to separate sheets in an Excel workbook based on "location_code_drawing type"
    output_file = "drawings_register.xlsx"
    save_drawings_to_excel(drawings_register, output_file)
    print("Saved updated drawings register to Excel file:", output_file)

    # Create a timestamped text file in the root folder
    create_timestamped_file()

    print("Finished processing.")

if __name__ == "__main__":
    main()